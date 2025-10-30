"""Utilities for working with IdealPOS exports."""

from __future__ import annotations

import re
from dataclasses import dataclass
from typing import Iterable, List, Sequence

from app.utils.numeric import coerce_float


_DEPARTMENT_HEADER_RE = re.compile(r"^(?P<code>\d+)\s+(?P<name>.+)$")


@dataclass
class DepartmentProductRow:
    """Normalized representation of a product row from a department export."""

    product_code: str | None
    product_name: str
    normalized_name: str
    quantity: float
    unit_price: float | None
    net_total: float | None
    row_number: int


@dataclass
class DepartmentBucket:
    """Collection of products that belong to a department bucket."""

    gl_code: str | None
    department_name: str
    rows: list[DepartmentProductRow]


@dataclass
class DepartmentSalesForecast:
    """Parsed IdealPOS export grouped into department buckets."""

    departments: list[DepartmentBucket]
    warnings: list[str]


def normalize_pos_alias(value: str) -> str:
    """Return a normalized representation of a POS product/location alias."""

    if not value:
        return ""
    value = value.strip().lower()
    value = re.sub(r"[^a-z0-9]+", " ", value)
    return re.sub(r"\s+", " ", value).strip()


_CURRENCY_PREFIX_RE = re.compile(
    r"^(?:[A-Za-z]{1,3}\s*\$?|[$€£¥₩₽])\s*",
    re.IGNORECASE,
)

_CURRENCY_SYMBOLS_RE = re.compile(r"[$€£¥₩₽]")


def parse_terminal_sales_number(value) -> float | None:
    """Best-effort conversion of spreadsheet values to ``float``."""

    if value is None:
        return None
    if isinstance(value, (int, float)):
        return float(value)
    try:
        text = str(value).replace("\u00A0", " ")
        cleaned = text.strip()
        if not cleaned:
            return None
        sign = ""
        if cleaned[0] in "+-":
            sign = cleaned[0]
            cleaned = cleaned[1:].lstrip()
        while True:
            stripped = _CURRENCY_PREFIX_RE.sub("", cleaned, count=1)
            if stripped == cleaned:
                break
            cleaned = stripped.lstrip()
        cleaned = (sign + cleaned).strip()
        if not cleaned:
            return None
        cleaned = _CURRENCY_SYMBOLS_RE.sub("", cleaned)
        cleaned = cleaned.replace(",", "")
        match = re.match(r"[-+]?\d*\.?\d+(?:[eE][-+]?\d+)?", cleaned)
        if match:
            return float(match.group(0))
        return float(cleaned)
    except (TypeError, ValueError):
        return None


def _is_effectively_zero(value: float | None) -> bool:
    if value is None:
        return True
    try:
        return abs(value) < 1e-9
    except (TypeError, ValueError):
        return True


def derive_terminal_sales_quantity(
    quantity: float | None,
    *,
    price: float | None = None,
    amount: float | None = None,
    net_including_tax_total: float | None = None,
    discounts_total: float | None = None,
) -> float | None:
    """Return a usable quantity, inferring it from totals when required."""

    if quantity is not None and not _is_effectively_zero(quantity):
        return float(quantity)

    if price in (None, 0.0):
        return quantity

    base_amount = amount
    if _is_effectively_zero(base_amount) and net_including_tax_total is not None:
        candidate = net_including_tax_total + (discounts_total or 0.0)
        if not _is_effectively_zero(candidate):
            base_amount = candidate

    if _is_effectively_zero(base_amount):
        return quantity

    try:
        inferred = float(base_amount) / float(price)
    except (TypeError, ValueError, ZeroDivisionError):
        return quantity

    if _is_effectively_zero(inferred):
        return quantity

    return inferred


def terminal_sales_cell_is_blank(value) -> bool:
    """Return ``True`` when an Excel cell should be treated as empty."""

    if value is None:
        return True
    if isinstance(value, str):
        return not value.strip()
    try:
        return float(value) == 0.0
    except (TypeError, ValueError):
        return False


def extract_terminal_sales_location(row: Sequence[object]) -> str | None:
    """Extract a location header from a parsed Excel row."""

    if not row:
        return None

    first = row[0]
    if not isinstance(first, str):
        return None

    candidate = first.strip()
    if not candidate:
        return None

    if all(terminal_sales_cell_is_blank(cell) for cell in row[1:]):
        return candidate

    return None


def group_terminal_sales_rows(row_data: Iterable[dict]) -> dict[str, dict]:
    """Group terminal sales rows by location and aggregate totals."""

    grouped: dict[str, dict] = {}
    for entry in row_data:
        loc = entry["location"]
        prod = entry.get("product")
        is_location_total = bool(entry.get("is_location_total"))
        qty_value = coerce_float(entry.get("quantity"))
        qty = qty_value if qty_value is not None else 0.0
        price = coerce_float(entry.get("price"))
        amount = coerce_float(entry.get("amount"))
        loc_entry = grouped.setdefault(
            loc,
            {
                "products": {},
                "total": 0.0,
                "total_amount": 0.0,
                "net_including_tax_total": 0.0,
                "discount_total": 0.0,
                "_has_net_including_tax_total": False,
                "_has_discount_total": False,
                "_raw_amount_total": 0.0,
                "_quantity_override": None,
                "_amount_override": None,
            },
        )
        if not is_location_total and prod:
            product_entry = loc_entry["products"].setdefault(
                prod,
                {
                    "quantity": 0.0,
                    "prices": [],
                    "amount": 0.0,
                },
            )
            product_entry["quantity"] += qty
            if price is not None:
                product_entry["prices"].append(price)
            if amount is not None:
                product_entry["amount"] += amount
                loc_entry["_raw_amount_total"] += amount
            loc_entry["total"] += qty
        elif is_location_total:
            if qty_value is not None:
                loc_entry["_quantity_override"] = qty_value
            if amount is not None:
                loc_entry["_amount_override"] = amount
        net_including_total = coerce_float(entry.get("net_including_tax_total"))
        if net_including_total is not None:
            loc_entry["net_including_tax_total"] += net_including_total
            loc_entry["_has_net_including_tax_total"] = True
        discount_total = coerce_float(entry.get("discount_total"))
        if discount_total is not None:
            loc_entry["discount_total"] += discount_total
            loc_entry["_has_discount_total"] = True
    for data in grouped.values():
        if not data["_has_net_including_tax_total"]:
            data["net_including_tax_total"] = None
        if not data["_has_discount_total"]:
            data["discount_total"] = None

        quantity_override = data.get("_quantity_override")
        if quantity_override is not None:
            data["total"] = quantity_override

        amount_override = data.get("_amount_override")
        if amount_override is not None:
            data["total_amount"] = amount_override
            continue

        net_total = data.get("net_including_tax_total")
        if net_total is not None:
            discount_total = data.get("discount_total") or 0.0
            data["total_amount"] = net_total + discount_total
        else:
            raw_amount_total = data.get("_raw_amount_total", 0.0) or 0.0
            data["total_amount"] = raw_amount_total
    return grouped


def iter_pos_excel_rows(filepath: str, extension: str) -> Iterable[List[object]]:
    """Yield rows from an IdealPOS export based on the file extension."""

    ext = extension.lower()
    if ext == ".xls":
        try:
            import xlrd  # type: ignore
        except ModuleNotFoundError:
            try:
                from app.vendor import xlrd  # type: ignore
            except ImportError:
                raise RuntimeError("legacy_xls_missing") from None
        try:
            book = xlrd.open_workbook(filepath)
        except Exception as exc:  # pragma: no cover - defensive
            raise RuntimeError("legacy_xls_error") from exc
        try:
            sheet = book.sheet_by_index(0)
        except IndexError as exc:  # pragma: no cover - defensive
            raise RuntimeError("legacy_xls_error") from exc
        try:
            for row_idx in range(sheet.nrows):
                yield [
                    sheet.cell_value(row_idx, col_idx)
                    for col_idx in range(sheet.ncols)
                ]
        finally:  # pragma: no branch - ensure resources freed
            try:
                book.release_resources()
            except AttributeError:
                pass
    elif ext == ".xlsx":
        try:
            from openpyxl import load_workbook
        except ImportError as exc:  # pragma: no cover - environment
            raise RuntimeError("xlsx_missing") from exc
        try:
            workbook = load_workbook(filepath, read_only=True, data_only=True)
        except Exception as exc:
            raise RuntimeError("xlsx_error") from exc
        try:
            sheet = workbook.active
            for row in sheet.iter_rows(values_only=True):
                yield list(row)
        finally:
            workbook.close()
    else:  # pragma: no cover - guarded by validators
        raise RuntimeError("unsupported_extension")


def parse_department_sales_forecast_rows(
    rows: Iterable[Sequence[object]],
) -> DepartmentSalesForecast:
    """Parse spreadsheet rows into department buckets suitable for forecasting."""

    departments: list[DepartmentBucket] = []
    warnings: list[str] = []
    current_bucket: DepartmentBucket | None = None

    for row_number, raw_row in enumerate(rows, start=1):
        row = list(raw_row)
        if not row:
            continue
        if all(terminal_sales_cell_is_blank(cell) for cell in row):
            continue
        header_name = extract_terminal_sales_location(row)
        if header_name:
            gl_code = None
            department_name = header_name
            match = _DEPARTMENT_HEADER_RE.match(header_name)
            if match:
                gl_code = match.group("code")
                department_name = match.group("name").strip()
            current_bucket = DepartmentBucket(
                gl_code=gl_code,
                department_name=department_name,
                rows=[],
            )
            departments.append(current_bucket)
            continue

        if current_bucket is None:
            warnings.append(
                "Encountered product rows before any department header; those rows were skipped."
            )
            continue

        product_cell = row[1] if len(row) > 1 else ""
        if not isinstance(product_cell, str):
            continue
        product_name = product_cell.strip()
        if not product_name:
            continue

        quantity = coerce_float(row[4] if len(row) > 4 else None, default=0.0) or 0.0
        unit_price = coerce_float(row[2] if len(row) > 2 else None)
        net_total = coerce_float(row[7] if len(row) > 7 else None)

        code_cell = row[0] if row else None
        product_code: str | None = None
        if isinstance(code_cell, str):
            product_code = code_cell.strip() or None
        elif isinstance(code_cell, (int, float)):
            if float(code_cell).is_integer():
                product_code = str(int(code_cell))
            else:
                product_code = str(code_cell)

        normalized_name = normalize_pos_alias(product_name)
        if not normalized_name:
            normalized_name = f"__unnamed_{row_number}"

        current_bucket.rows.append(
            DepartmentProductRow(
                product_code=product_code,
                product_name=product_name,
                normalized_name=normalized_name,
                quantity=quantity,
                unit_price=unit_price,
                net_total=net_total,
                row_number=row_number,
            )
        )

    return DepartmentSalesForecast(departments=departments, warnings=warnings)


def parse_department_sales_forecast(filepath: str, extension: str) -> DepartmentSalesForecast:
    """Parse an IdealPOS department sales export from disk."""

    rows = iter_pos_excel_rows(filepath, extension)
    return parse_department_sales_forecast_rows(rows)

