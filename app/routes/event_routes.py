import csv
import io
import json
import math
import os
import re
from collections import defaultdict
from datetime import datetime
from types import SimpleNamespace

from flask import (
    Blueprint,
    abort,
    current_app,
    flash,
    jsonify,
    make_response,
    redirect,
    render_template,
    request,
    url_for,
)
from flask_login import login_required
from werkzeug.utils import secure_filename

from app import db
from app.forms import (
    EVENT_TYPES,
    EventForm,
    EventLocationConfirmForm,
    EventLocationForm,
    UpdateOpeningCountsForm,
    ScanCountForm,
    TerminalSalesUploadForm,
)
from app.models import (
    Event,
    EventLocation,
    EventLocationTerminalSalesSummary,
    EventStandSheetItem,
    Item,
    Location,
    LocationStandItem,
    Product,
    TerminalSale,
    TerminalSaleProductAlias,
)
from app.utils.activity import log_activity
from app.utils.numeric import coerce_float
from app.utils.units import (
    DEFAULT_BASE_UNIT_CONVERSIONS,
    convert_quantity,
    convert_quantity_for_reporting,
    get_unit_label,
)
from itsdangerous import BadSignature, URLSafeSerializer

_STAND_SHEET_FIELDS = (
    "opening_count",
    "transferred_in",
    "transferred_out",
    "eaten",
    "spoiled",
    "closing_count",
)


def _conversion_mapping():
    """Return the configured reporting-unit conversions."""

    configured = current_app.config.get("BASE_UNIT_CONVERSIONS") or {}
    merged = dict(DEFAULT_BASE_UNIT_CONVERSIONS)
    merged.update(configured)
    return merged


def _convert_value_for_reporting(value, base_unit, conversions):
    """Convert a stored base-unit value for presentation."""

    if value is None or not base_unit:
        return value
    try:
        converted, _ = convert_quantity_for_reporting(
            float(value), base_unit, conversions
        )
    except (TypeError, ValueError):
        return value
    return converted


def _build_sheet_values(sheet, base_unit, conversions):
    """Return reporting-unit stand sheet values for display."""

    values = {}
    for field in _STAND_SHEET_FIELDS:
        raw = getattr(sheet, field, None) if sheet else None
        values[field] = (
            _convert_value_for_reporting(raw, base_unit, conversions)
            if raw is not None
            else None
        )
    return SimpleNamespace(**values)


def _build_stand_item_entry(
    *,
    item,
    expected=0.0,
    sales=0.0,
    sheet=None,
    recv_unit=None,
    trans_unit=None,
    conversions=None,
):
    """Assemble a stand-sheet entry enriched with reporting metadata."""

    conversions = conversions or _conversion_mapping()
    base_unit = item.base_unit
    report_unit = conversions.get(base_unit, base_unit)
    report_label = get_unit_label(report_unit)
    expected_display = _convert_value_for_reporting(expected, base_unit, conversions)
    sales_display = _convert_value_for_reporting(sales, base_unit, conversions)
    if sales_display is None:
        sales_display = 0.0
    return {
        "item": item,
        "expected": expected_display,
        "expected_base": expected,
        "sales": sales_display,
        "sales_base": sales,
        "sheet": sheet,
        "sheet_values": _build_sheet_values(sheet, base_unit, conversions),
        "base_unit": base_unit,
        "report_unit": report_unit,
        "report_unit_label": report_label,
        "recv_unit": recv_unit,
        "trans_unit": trans_unit,
    }


def _sync_event_location_opening_counts(event_location: EventLocation) -> int:
    """Ensure stand sheet opening counts mirror the location inventory."""

    inventory_records = LocationStandItem.query.filter_by(
        location_id=event_location.location_id
    ).all()
    if not inventory_records:
        return 0

    existing_sheets = {
        sheet.item_id: sheet
        for sheet in EventStandSheetItem.query.filter_by(
            event_location_id=event_location.id
        )
    }

    updated = 0
    for record in inventory_records:
        sheet = existing_sheets.get(record.item_id)
        if sheet is None:
            sheet = EventStandSheetItem(
                event_location_id=event_location.id,
                item_id=record.item_id,
            )
            db.session.add(sheet)
        sheet.opening_count = float(record.expected_count or 0.0)
        updated += 1

    return updated


def _convert_report_value_to_base(value, base_unit, report_unit):
    """Convert a reporting-unit form value back into the base unit."""

    if value is None:
        return 0.0
    if not base_unit or not report_unit or base_unit == report_unit:
        return value
    try:
        return convert_quantity(value, report_unit, base_unit)
    except (TypeError, ValueError):
        return value


event = Blueprint("event", __name__)


def _terminal_sales_serializer() -> URLSafeSerializer:
    secret_key = current_app.secret_key or current_app.config.get("SECRET_KEY")
    if not secret_key:
        raise RuntimeError("Application secret key is not configured.")
    return URLSafeSerializer(secret_key, salt="terminal-sales-resolution")


def _parse_date(value):
    if not value:
        return None
    try:
        return datetime.strptime(value, "%Y-%m-%d").date()
    except ValueError:
        return None


def _get_event_filters(source):
    return {
        "type": (source.get("type") or "").strip(),
        "name_contains": (source.get("name_contains") or "").strip(),
        "name_not_contains": (source.get("name_not_contains") or "").strip(),
        "start_date_from": (source.get("start_date_from") or "").strip(),
        "start_date_to": (source.get("start_date_to") or "").strip(),
        "end_date_from": (source.get("end_date_from") or "").strip(),
        "end_date_to": (source.get("end_date_to") or "").strip(),
        "closed_status": (source.get("closed_status") or "").strip(),
    }


def _apply_event_filters(query, filters):
    event_type = filters.get("type")
    if event_type:
        query = query.filter_by(event_type=event_type)

    name_contains = filters.get("name_contains")
    if name_contains:
        query = query.filter(Event.name.ilike(f"%{name_contains}%"))

    name_not_contains = filters.get("name_not_contains")
    if name_not_contains:
        query = query.filter(~Event.name.ilike(f"%{name_not_contains}%"))

    start_date_from = _parse_date(filters.get("start_date_from"))
    if start_date_from:
        query = query.filter(Event.start_date >= start_date_from)

    start_date_to = _parse_date(filters.get("start_date_to"))
    if start_date_to:
        query = query.filter(Event.start_date <= start_date_to)

    end_date_from = _parse_date(filters.get("end_date_from"))
    if end_date_from:
        query = query.filter(Event.end_date >= end_date_from)

    end_date_to = _parse_date(filters.get("end_date_to"))
    if end_date_to:
        query = query.filter(Event.end_date <= end_date_to)

    closed_status = filters.get("closed_status")
    if closed_status == "open":
        query = query.filter(Event.closed.is_(False))
    elif closed_status == "closed":
        query = query.filter(Event.closed.is_(True))

    return query


@event.route("/events")
@login_required
def view_events():
    filters = _get_event_filters(request.args)
    query = _apply_event_filters(Event.query, filters)
    events = query.all()
    create_form = EventForm()
    return render_template(
        "events/view_events.html",
        events=events,
        event_types=EVENT_TYPES,
        type_labels=dict(EVENT_TYPES),
        create_form=create_form,
        filter_values=filters,
    )


@event.route("/events/create", methods=["GET", "POST"])
@login_required
def create_event():
    form = EventForm()
    if form.validate_on_submit():
        ev = Event(
            name=form.name.data,
            start_date=form.start_date.data,
            end_date=form.end_date.data,
            event_type=form.event_type.data,
            estimated_sales=form.estimated_sales.data,
        )
        db.session.add(ev)
        db.session.commit()
        log_activity(f"Created event {ev.id}")
        flash("Event created")
        return redirect(url_for("event.view_events"))
    return render_template("events/create_event.html", form=form)


@event.route("/events/filter", methods=["POST"])
@login_required
def filter_events_ajax():
    filters = _get_event_filters(request.form)
    events = _apply_event_filters(Event.query, filters).all()
    return render_template(
        "events/_events_table.html",
        events=events,
        type_labels=dict(EVENT_TYPES),
    )


@event.route("/events/create/ajax", methods=["POST"])
@login_required
def create_event_ajax():
    form = EventForm()
    if form.validate_on_submit():
        ev = Event(
            name=form.name.data,
            start_date=form.start_date.data,
            end_date=form.end_date.data,
            event_type=form.event_type.data,
            estimated_sales=form.estimated_sales.data,
        )
        db.session.add(ev)
        db.session.commit()
        log_activity(f"Created event {ev.id}")
        return render_template(
            "events/_event_row.html", e=ev, type_labels=dict(EVENT_TYPES)
        )
    response = {"errors": form.errors or {"form": ["Invalid data submitted."]}}
    return jsonify(response), 400


@event.route("/events/<int:event_id>/edit", methods=["GET", "POST"])
@login_required
def edit_event(event_id):
    ev = db.session.get(Event, event_id)
    if ev is None:
        abort(404)
    form = EventForm(obj=ev)
    if form.validate_on_submit():
        ev.name = form.name.data
        ev.start_date = form.start_date.data
        ev.end_date = form.end_date.data
        ev.event_type = form.event_type.data
        ev.estimated_sales = form.estimated_sales.data
        db.session.commit()
        log_activity(f"Edited event {ev.id}")
        flash("Event updated")
        return redirect(url_for("event.view_events"))
    return render_template("events/edit_event.html", form=form, event=ev)


@event.route("/events/<int:event_id>/delete")
@login_required
def delete_event(event_id):
    ev = db.session.get(Event, event_id)
    if ev is None:
        abort(404)
    event_id = ev.id
    db.session.delete(ev)
    db.session.commit()
    log_activity(f"Deleted event {event_id}")
    flash("Event deleted")
    return redirect(url_for("event.view_events"))


@event.route("/events/<int:event_id>")
@login_required
def view_event(event_id):
    ev = db.session.get(Event, event_id)
    if ev is None:
        abort(404)
    type_labels = dict(EVENT_TYPES)
    opening_form = UpdateOpeningCountsForm()
    opening_form.location_ids.choices = [
        (el.id, el.location.name)
        for el in ev.locations
        if el.location is not None
    ]
    return render_template(
        "events/view_event.html",
        event=ev,
        event_type_label=type_labels.get(ev.event_type, ev.event_type),
        opening_form=opening_form,
    )


@event.route(
    "/events/<int:event_id>/update_opening_counts", methods=["POST"]
)
@login_required
def update_opening_counts(event_id):
    ev = db.session.get(Event, event_id)
    if ev is None:
        abort(404)

    form = UpdateOpeningCountsForm()
    event_locations = (
        EventLocation.query.filter_by(event_id=event_id)
        .join(Location)
        .order_by(Location.name)
        .all()
    )
    form.location_ids.choices = [
        (el.id, el.location.name) for el in event_locations
    ]

    if not form.validate_on_submit():
        flash("Unable to update opening counts. Please try again.", "warning")
        return redirect(url_for("event.view_event", event_id=event_id))

    if ev.closed:
        flash("This event is closed and opening counts cannot be updated.", "warning")
        return redirect(url_for("event.view_event", event_id=event_id))

    selected_ids = form.location_ids.data or []
    if not selected_ids:
        flash("Select at least one location to update opening counts.", "warning")
        return redirect(url_for("event.view_event", event_id=event_id))

    location_map = {el.id: el for el in event_locations}
    updated_names = []
    skipped_names = []
    for el_id in selected_ids:
        el = location_map.get(el_id)
        if el is None:
            continue
        if el.confirmed:
            skipped_names.append(el.location.name)
            continue
        _sync_event_location_opening_counts(el)
        updated_names.append(el.location.name)

    if not updated_names:
        if skipped_names:
            flash(
                "The selected locations are already confirmed and cannot be updated.",
                "warning",
            )
        else:
            flash("No matching locations were found to update.", "warning")
        return redirect(url_for("event.view_event", event_id=event_id))

    db.session.commit()

    log_activity(
        "Updated opening counts for event %s locations: %s"
        % (event_id, ", ".join(updated_names))
    )

    message = "Opening counts updated for: %s" % ", ".join(updated_names)
    if skipped_names:
        message += ". Skipped confirmed locations: %s" % ", ".join(skipped_names)
    flash(message, "success")

    return redirect(url_for("event.view_event", event_id=event_id))


@event.route("/events/<int:event_id>/add_location", methods=["GET", "POST"])
@login_required
def add_location(event_id):
    ev = db.session.get(Event, event_id)
    if ev is None:
        abort(404)
    form = EventLocationForm(event_id=event_id)
    if not form.location_id.choices:
        flash("All available locations have already been assigned to this event.")
        return redirect(url_for("event.view_event", event_id=event_id))
    if form.validate_on_submit():
        selected_ids = form.location_id.data
        event_locations = []
        for location_id in selected_ids:
            event_location = EventLocation(
                event_id=event_id, location_id=location_id
            )
            db.session.add(event_location)
            event_locations.append(event_location)
        if event_locations:
            db.session.flush()
            for event_location in event_locations:
                _sync_event_location_opening_counts(event_location)
        db.session.commit()
        location_names = []
        for location_id in selected_ids:
            location = db.session.get(Location, location_id)
            location_names.append(location.name if location else str(location_id))
        location_list = ", ".join(location_names)
        log_activity(
            f"Assigned locations {location_list} to event {event_id}"
        )
        flash(
            "Locations assigned"
            if len(event_locations) > 1
            else "Location assigned"
        )
        return redirect(url_for("event.view_event", event_id=event_id))
    return render_template("events/add_location.html", form=form, event=ev)


@event.route(
    "/events/<int:event_id>/locations/<int:el_id>/sales/add",
    methods=["GET", "POST"],
)
@login_required
def add_terminal_sale(event_id, el_id):
    el = db.session.get(EventLocation, el_id)
    if el is None or el.event_id != event_id:
        abort(404)
    if el.confirmed or el.event.closed:
        flash("This location is closed and cannot accept new sales.")
        return redirect(url_for("event.view_event", event_id=event_id))
    if request.method == "POST":
        updated = False
        for product in el.location.products:
            qty = request.form.get(f"qty_{product.id}")
            try:
                amount = float(qty) if qty else 0
            except ValueError:
                amount = 0

            sale = TerminalSale.query.filter_by(
                event_location_id=el_id, product_id=product.id
            ).first()

            if amount:
                if sale:
                    if sale.quantity != amount:
                        sale.quantity = amount
                        updated = True
                else:
                    sale = TerminalSale(
                        event_location_id=el_id,
                        product_id=product.id,
                        quantity=amount,
                        sold_at=datetime.utcnow(),
                    )
                    db.session.add(sale)
                    updated = True
            elif sale:
                db.session.delete(sale)
                updated = True

        db.session.commit()
        if updated:
            log_activity(f"Updated terminal sales for event location {el_id}")
        flash("Sales recorded")
        return redirect(url_for("event.view_event", event_id=event_id))

    existing_sales = {s.product_id: s.quantity for s in el.terminal_sales}
    return render_template(
        "events/add_terminal_sales.html",
        event_location=el,
        existing_sales=existing_sales,
    )


def _wants_json_response() -> bool:
    """Return True when the current request prefers a JSON response."""

    if request.is_json:
        return True
    accept_mimetypes = request.accept_mimetypes
    return (
        accept_mimetypes["application/json"]
        > accept_mimetypes["text/html"]
    )


def _serialize_scan_totals(event_location: EventLocation):
    """Return the location and summaries of counted items."""

    location, stand_items = _get_stand_items(
        event_location.location_id, event_location.event_id
    )
    totals = []
    seen_item_ids = set()

    for entry in stand_items:
        item = entry["item"]
        sheet = entry.get("sheet")
        counted = float(sheet.closing_count or 0.0) if sheet else 0.0
        totals.append(
            {
                "item_id": item.id,
                "name": item.name,
                "upc": item.upc,
                "expected": float(entry.get("expected") or 0.0),
                "counted": counted,
                "base_unit": item.base_unit,
            }
        )
        seen_item_ids.add(item.id)

    for sheet in event_location.stand_sheet_items:
        if sheet.item_id in seen_item_ids:
            continue
        item = sheet.item
        totals.append(
            {
                "item_id": item.id,
                "name": item.name,
                "upc": item.upc,
                "expected": 0.0,
                "counted": float(sheet.closing_count or 0.0),
                "base_unit": item.base_unit,
            }
        )

    totals.sort(key=lambda record: record["name"].lower())
    return location, totals


@event.route(
    "/events/<int:event_id>/locations/<int:location_id>/scan_counts",
    methods=["GET", "POST"],
)
@login_required
def scan_counts(event_id, location_id):
    ev = db.session.get(Event, event_id)
    if ev is None:
        abort(404)
    if ev.event_type != "inventory":
        abort(404)

    el = EventLocation.query.filter_by(
        event_id=event_id, location_id=location_id
    ).first()
    if el is None:
        abort(404)

    wants_json = _wants_json_response()

    if ev.closed:
        if wants_json:
            return (
                jsonify(
                    success=False, error="This event is closed to updates."
                ),
                403,
            )
        abort(403, description="This event is closed to updates.")

    form = ScanCountForm()
    if form.quantity.data is None:
        form.quantity.data = 1

    if request.method == "GET" and wants_json:
        location, totals = _serialize_scan_totals(el)
        return jsonify(
            success=True,
            location={"id": location.id, "name": location.name},
            totals=totals,
        )

    if request.method == "POST":
        if wants_json:
            payload = request.get_json(silent=True) or {}
            upc = (payload.get("upc") or "").strip()
            raw_quantity = payload.get("quantity", 1)
            try:
                quantity = float(raw_quantity)
            except (TypeError, ValueError):
                quantity = None

            if not upc:
                return (
                    jsonify(success=False, error="A UPC value is required."),
                    400,
                )
            if quantity is None:
                return (
                    jsonify(
                        success=False,
                        error="Quantity must be a numeric value.",
                    ),
                    400,
                )
        else:
            if not form.validate_on_submit():
                location, totals = _serialize_scan_totals(el)
                return (
                    render_template(
                        "events/scan_count.html",
                        event=ev,
                        location=location,
                        form=form,
                        totals=totals,
                    ),
                    400,
                )
            upc = (form.upc.data or "").strip()
            quantity = float(form.quantity.data or 0)

        item = Item.query.filter_by(upc=upc).first()
        if item is None:
            if wants_json:
                return (
                    jsonify(
                        success=False,
                        error=f"No item found for UPC {upc}.",
                    ),
                    404,
                )
            flash(f"No item found for UPC {upc}.", "danger")
            location, totals = _serialize_scan_totals(el)
            return (
                render_template(
                    "events/scan_count.html",
                    event=ev,
                    location=location,
                    form=form,
                    totals=totals,
                ),
                404,
            )

        sheet = EventStandSheetItem.query.filter_by(
            event_location_id=el.id, item_id=item.id
        ).first()
        if sheet is None:
            sheet = EventStandSheetItem(
                event_location_id=el.id, item_id=item.id
            )
            db.session.add(sheet)

        sheet.transferred_out = (sheet.transferred_out or 0.0) + quantity
        sheet.closing_count = (sheet.closing_count or 0.0) + quantity
        db.session.commit()
        log_activity(
            f"Recorded scan count for event {event_id}"
            f" location {location_id} item {item.id}"
        )

        location, totals = _serialize_scan_totals(el)

        if wants_json:
            return jsonify(
                success=True,
                item={
                    "id": item.id,
                    "name": item.name,
                    "upc": item.upc,
                    "quantity": quantity,
                    "total": float(sheet.transferred_out or 0.0),
                    "base_unit": item.base_unit,
                },
                totals=totals,
            )

        flash(
            f"Recorded {quantity:g} {item.base_unit} for {item.name}.",
            "success",
        )
        return redirect(
            url_for(
                "event.scan_counts",
                event_id=event_id,
                location_id=location_id,
            )
        )

    location, totals = _serialize_scan_totals(el)
    return render_template(
        "events/scan_count.html",
        event=ev,
        location=location,
        form=form,
        totals=totals,
    )


@event.route("/events/<int:event_id>/sales/upload", methods=["GET", "POST"])
@login_required
def upload_terminal_sales(event_id):
    ev = db.session.get(Event, event_id)
    if ev is None:
        abort(404)

    form = TerminalSalesUploadForm()
    open_locations = [
        el
        for el in ev.locations
        if not el.confirmed and not ev.closed
    ]
    mapping_payload = None
    mapping_filename = None
    sales_summary: dict[str, dict] = {}
    sales_location_names: list[str] = []
    default_mapping: dict[int, str] = {}
    unresolved_products: list[dict] = []
    resolution_errors: list[str] = []
    product_resolution_required = False
    product_choices: list[Product] = []
    price_discrepancies: dict[str, list[dict]] = {}
    menu_mismatches: dict[str, list[dict]] = {}
    warnings_required = False
    warnings_acknowledged = False

    def _normalize_product_name(value: str) -> str:
        if not value:
            return ""
        value = value.strip().lower()
        value = re.sub(r"[^a-z0-9]+", " ", value)
        return re.sub(r"\s+", " ", value).strip()

    def _to_float(value):
        if value is None:
            return None
        if isinstance(value, (int, float)):
            return float(value)
        try:
            cleaned = str(value).strip().replace("$", "")
            if not cleaned:
                return None
            return float(cleaned)
        except (TypeError, ValueError):
            return None

    def _group_rows(row_data):
        grouped: dict[str, dict] = {}
        for entry in row_data:
            loc = entry["location"]
            prod = entry["product"]
            qty = float(entry.get("quantity", 0.0))
            price = entry.get("price")
            amount = entry.get("amount")
            loc_entry = grouped.setdefault(
                loc,
                {
                    "products": {},
                    "total": 0.0,
                    "total_amount": 0.0,
                    "net_including_tax_total": 0.0,
                    "discount_total": 0.0,
                    "_has_net_including_tax_total": False,
                    "_has_discount_total": False,
                    "_raw_amount_total": 0.0,
                },
            )
            product_entry = loc_entry["products"].setdefault(
                prod,
                {
                    "quantity": 0.0,
                    "prices": [],
                    "amount": 0.0,
                },
            )
            product_entry["quantity"] += qty
            if price is not None:
                product_entry["prices"].append(price)
            if amount is not None:
                product_entry["amount"] += amount
                loc_entry["_raw_amount_total"] += amount
            loc_entry["total"] += qty
            net_including_total = entry.get("net_including_tax_total")
            if net_including_total is not None:
                loc_entry["net_including_tax_total"] += net_including_total
                loc_entry["_has_net_including_tax_total"] = True
            discount_total = entry.get("discount_total")
            if discount_total is not None:
                loc_entry["discount_total"] += discount_total
                loc_entry["_has_discount_total"] = True

        for loc_entry in grouped.values():
            for product_entry in loc_entry["products"].values():
                if product_entry["prices"]:
                    unique_prices = sorted({round(p, 4) for p in product_entry["prices"]})
                    product_entry["prices"] = unique_prices
            has_net = loc_entry.pop("_has_net_including_tax_total", False)
            has_discount = loc_entry.pop("_has_discount_total", False)
            raw_total = loc_entry.pop("_raw_amount_total", 0.0)
            if has_net or has_discount:
                net_total = loc_entry["net_including_tax_total"] if has_net else 0.0
                discount_total = loc_entry["discount_total"] if has_discount else 0.0
                loc_entry["total_amount"] = net_total + discount_total
            else:
                loc_entry["total_amount"] = raw_total
        return grouped

    def _derive_price_map(summary: dict[str, dict]) -> dict[str, float | None]:
        """Build a mapping of product names to representative sale prices."""

        price_map: dict[str, float | None] = {}
        for data in summary.values():
            products = data.get("products", {})
            for name, details in products.items():
                if name in price_map and price_map[name] is not None:
                    continue
                candidate = None
                prices = details.get("prices") or []
                for price in prices:
                    if price is None:
                        continue
                    try:
                        candidate = float(price)
                    except (TypeError, ValueError):
                        continue
                    else:
                        break
                if candidate is None:
                    amount = details.get("amount")
                    quantity = details.get("quantity")
                    try:
                        if amount is not None and quantity:
                            candidate = float(amount) / float(quantity)
                    except (TypeError, ValueError, ZeroDivisionError):
                        candidate = None
                price_map.setdefault(name, candidate)
        return price_map

    def _prices_match(file_price: float, app_price: float) -> bool:
        try:
            return math.isclose(float(file_price), float(app_price), abs_tol=0.01)
        except (TypeError, ValueError):
            return False

    def _ensure_location_items(location_obj: Location, product_obj: Product) -> None:
        for recipe_item in product_obj.recipe_items:
            if not recipe_item.countable:
                continue
            record = LocationStandItem.query.filter_by(
                location_id=location_obj.id, item_id=recipe_item.item_id
            ).first()
            if record is None:
                db.session.add(
                    LocationStandItem(
                        location_id=location_obj.id,
                        item_id=recipe_item.item_id,
                        expected_count=0,
                        purchase_gl_code_id=recipe_item.item.purchase_gl_code_id,
                    )
                )

    def _apply_pending_sales(
        pending_sales: list[dict],
        pending_totals: list[dict] | None = None,
    ) -> set[str]:
        updated_locations: set[str] = set()
        totals_map: dict[int, dict] = {}
        if pending_totals:
            for entry in pending_totals:
                try:
                    el_id = int(entry.get("event_location_id"))
                except (TypeError, ValueError):
                    continue
                totals_map[el_id] = entry
        for entry in pending_sales:
            event_location_id = entry.get("event_location_id")
            product_id = entry.get("product_id")
            quantity_value = entry.get("quantity", 0.0)
            if not event_location_id or not product_id:
                continue
            event_location = db.session.get(EventLocation, event_location_id)
            product = db.session.get(Product, product_id)
            if event_location is None or product is None:
                continue
            sale = TerminalSale.query.filter_by(
                event_location_id=event_location.id, product_id=product.id
            ).first()
            if sale:
                sale.quantity = quantity_value
            else:
                db.session.add(
                    TerminalSale(
                        event_location_id=event_location.id,
                        product_id=product.id,
                        quantity=quantity_value,
                        sold_at=datetime.utcnow(),
                    )
                )
            updated_locations.add(event_location.location.name)
        if totals_map:
            for el_id, data in totals_map.items():
                summary = EventLocationTerminalSalesSummary.query.filter_by(
                    event_location_id=el_id
                ).first()
                if summary is None:
                    summary = EventLocationTerminalSalesSummary(
                        event_location_id=el_id
                    )
                    db.session.add(summary)
                summary.source_location = data.get("source_location")
                summary.total_quantity = data.get("total_quantity")
                summary.total_amount = data.get("total_amount")
        return updated_locations

    def _apply_resolution_actions(issue_state: dict) -> tuple[list[str], list[str]]:
        price_updates: list[str] = []
        menu_updates: list[str] = []
        for location_issue in issue_state.get("queue", []):
            event_location_id = location_issue.get("event_location_id")
            event_location = None
            if event_location_id:
                event_location = db.session.get(EventLocation, event_location_id)
            for issue in location_issue.get("price_issues", []):
                if issue.get("resolution") != "update":
                    continue
                product_id = issue.get("product_id")
                new_price = issue.get("target_price")
                if product_id is None or new_price is None:
                    continue
                product = db.session.get(Product, product_id)
                if product is None:
                    continue
                product.price = new_price
                price_updates.append(product.name)
            if event_location is None:
                continue
            location_obj = event_location.location
            if location_obj is None:
                continue
            for issue in location_issue.get("menu_issues", []):
                if issue.get("resolution") != "add":
                    continue
                product_id = issue.get("product_id")
                if product_id is None:
                    continue
                product = db.session.get(Product, product_id)
                if product is None:
                    continue
                if product not in location_obj.products:
                    location_obj.products.append(product)
                if (
                    location_obj.current_menu
                    and product not in location_obj.current_menu.products
                ):
                    location_obj.current_menu.products.append(product)
                _ensure_location_items(location_obj, product)
                menu_updates.append(f"{product.name} @ {location_obj.name}")
        return price_updates, menu_updates

    if request.method == "POST":
        step = request.form.get("step")
        if step == "resolve":
            payload = request.form.get("payload")
            state_token = request.form.get("state_token")
            mapping_filename = request.form.get("mapping_filename")
            if not state_token:
                flash("Unable to continue the resolution process.", "danger")
                return redirect(url_for("event.upload_terminal_sales", event_id=event_id))
            serializer = _terminal_sales_serializer()
            try:
                state_data = serializer.loads(state_token)
            except BadSignature:
                flash("The resolution data could not be verified.", "danger")
                return redirect(url_for("event.upload_terminal_sales", event_id=event_id))

            queue: list[dict] = state_data.get("queue") or []
            pending_sales: list[dict] = state_data.get("pending_sales") or []
            pending_totals: list[dict] = state_data.get("pending_totals") or []
            selected_locations: list[str] = state_data.get("selected_locations") or []
            issue_index = state_data.get("issue_index", 0)
            action = request.form.get("action", "")

            if issue_index < 0:
                issue_index = 0
            if issue_index > len(queue):
                issue_index = len(queue)

            if queue and issue_index < len(queue):
                current_issue = queue[issue_index]
            else:
                current_issue = None

            error_messages: list[str] = []

            if action.startswith("price:") and current_issue:
                parts = action.split(":", 2)
                if len(parts) == 3:
                    _, product_id_raw, resolution_value = parts
                    try:
                        product_id_int = int(product_id_raw)
                    except (TypeError, ValueError):
                        error_messages.append("Invalid price resolution request.")
                    else:
                        for issue in current_issue.get("price_issues", []):
                            if issue.get("product_id") == product_id_int:
                                if resolution_value == "update":
                                    issue["resolution"] = "update"
                                elif resolution_value == "skip":
                                    issue["resolution"] = "skip"
                                break
                else:
                    error_messages.append("Invalid price resolution request.")
            elif action.startswith("menu:") and current_issue:
                parts = action.split(":", 2)
                if len(parts) == 3:
                    _, product_id_raw, resolution_value = parts
                    try:
                        product_id_int = int(product_id_raw)
                    except (TypeError, ValueError):
                        error_messages.append("Invalid menu resolution request.")
                    else:
                        for issue in current_issue.get("menu_issues", []):
                            if issue.get("product_id") == product_id_int:
                                if resolution_value == "add":
                                    issue["resolution"] = "add"
                                elif resolution_value == "skip":
                                    issue["resolution"] = "skip"
                                break
                else:
                    error_messages.append("Invalid menu resolution request.")
            elif action == "next_location":
                if current_issue:
                    unresolved = [
                        issue
                        for issue in current_issue.get("price_issues", [])
                        if issue.get("resolution") is None
                    ]
                    unresolved.extend(
                        issue
                        for issue in current_issue.get("menu_issues", [])
                        if issue.get("resolution") is None
                    )
                    if unresolved:
                        error_messages.append(
                            "Resolve all issues for this location before continuing."
                        )
                    else:
                        issue_index += 1
            elif action == "finish":
                unresolved_overall = []
                for location_issue in queue:
                    unresolved_overall.extend(
                        issue
                        for issue in location_issue.get("price_issues", [])
                        if issue.get("resolution") is None
                    )
                    unresolved_overall.extend(
                        issue
                        for issue in location_issue.get("menu_issues", [])
                        if issue.get("resolution") is None
                    )
                if unresolved_overall:
                    error_messages.append(
                        "Resolve all issues before finishing the import."
                    )
                else:
                    issue_index = len(queue)

            if queue and issue_index < len(queue):
                current_issue = queue[issue_index]
            else:
                current_issue = None

            if issue_index >= len(queue):
                updated_locations = _apply_pending_sales(
                    pending_sales, pending_totals
                )
                price_updates, menu_updates = _apply_resolution_actions(
                    {"queue": queue}
                )
                if updated_locations or price_updates or menu_updates:
                    db.session.commit()
                    log_activity(
                        "Uploaded terminal sales for event "
                        f"{event_id} from {mapping_filename or 'uploaded file'}"
                    )
                    success_parts: list[str] = []
                    if updated_locations:
                        success_parts.append(
                            "Terminal sales were imported for: "
                            + ", ".join(sorted(updated_locations))
                        )
                    if price_updates:
                        success_parts.append(
                            "Updated product prices: " + ", ".join(sorted(set(price_updates)))
                        )
                    if menu_updates:
                        success_parts.append(
                            "Added products to menus: " + ", ".join(sorted(set(menu_updates)))
                        )
                    flash(" ".join(success_parts), "success")
                else:
                    flash(
                        "No event locations were linked to the uploaded sales data.",
                        "warning",
                    )
                return redirect(url_for("event.view_event", event_id=event_id))

            if error_messages:
                for message in error_messages:
                    flash(message, "danger")

            state_data["queue"] = queue
            state_data["pending_sales"] = pending_sales
            state_data["pending_totals"] = pending_totals
            state_data["selected_locations"] = selected_locations
            state_data["issue_index"] = issue_index
            state_token = serializer.dumps(state_data)

            total_locations = len(queue)
            return render_template(
                "events/upload_terminal_sales.html",
                form=form,
                event=ev,
                open_locations=open_locations,
                mapping_payload=payload,
                mapping_filename=mapping_filename,
                sales_summary={},
                sales_location_names=[],
                default_mapping={},
                unresolved_products=[],
                product_choices=[],
                resolution_errors=[],
                product_resolution_required=False,
                price_discrepancies={},
                menu_mismatches={},
                warnings_required=False,
                warnings_acknowledged=False,
                state_token=state_token,
                issue_index=issue_index,
                current_issue=current_issue,
                remaining_locations=len(queue) - issue_index - 1,
                selected_locations=selected_locations,
                issue_total=total_locations,
            )

        elif step == "map":
            payload = request.form.get("payload")
            if not payload:
                flash("Unable to process the uploaded sales data.", "danger")
                return redirect(url_for("event.upload_terminal_sales", event_id=event_id))
            try:
                payload_data = json.loads(payload)
            except (TypeError, ValueError):
                flash("The uploaded sales data is invalid.", "danger")
                return redirect(url_for("event.upload_terminal_sales", event_id=event_id))

            rows = payload_data.get("rows", [])
            mapping_filename = payload_data.get("filename")
            if not rows:
                flash("No sales records were found in the uploaded file.", "warning")
                return redirect(url_for("event.upload_terminal_sales", event_id=event_id))

            sales_summary = _group_rows(rows)
            product_price_lookup = _derive_price_map(sales_summary)

            product_names = {
                product_name
                for data in sales_summary.values()
                for product_name in data["products"].keys()
            }
            product_lookup: dict[str, Product] = {}
            normalized_lookup = {
                name: _normalize_product_name(name) for name in product_names
            }

            if product_names:
                product_lookup.update(
                    {
                        p.name: p
                        for p in Product.query.filter(
                            Product.name.in_(product_names)
                        ).all()
                    }
                )

                normalized_values = [
                    norm for norm in normalized_lookup.values() if norm
                ]
                alias_lookup: dict[str, TerminalSaleProductAlias] = {}
                if normalized_values:
                    alias_rows = (
                        TerminalSaleProductAlias.query.filter(
                            TerminalSaleProductAlias.normalized_name.in_(
                                normalized_values
                            )
                        ).all()
                    )
                    alias_lookup = {
                        alias.normalized_name: alias for alias in alias_rows
                    }
                    for original_name, normalized in normalized_lookup.items():
                        if (
                            normalized
                            and original_name not in product_lookup
                            and normalized in alias_lookup
                        ):
                            product_lookup[original_name] = alias_lookup[
                                normalized
                            ].product
            else:
                alias_lookup = {}

            unmatched_names = [
                name for name in product_names if name not in product_lookup
            ]

            if unmatched_names:
                product_resolution_required = True
                resolution_requested = request.form.get(
                    "product-resolution-step"
                ) == "1"
                if not product_choices:
                    product_choices = Product.query.order_by(Product.name).all()

                manual_mappings: dict[str, Product] = {}
                pending_creations: list[str] = []
                CREATE_SELECTION_VALUE = "__create__"
                for idx, original_name in enumerate(unmatched_names):
                    field_name = f"product-match-{idx}"
                    selected_value = request.form.get(field_name)
                    selected_product = None
                    if selected_value:
                        if selected_value == CREATE_SELECTION_VALUE:
                            pending_creations.append(original_name)
                        else:
                            try:
                                product_id = int(selected_value)
                            except (TypeError, ValueError):
                                resolution_errors.append(
                                    f"Invalid product selection for {original_name}."
                                )
                            else:
                                selected_product = db.session.get(
                                    Product, product_id
                                )
                                if selected_product is None:
                                    resolution_errors.append(
                                        f"Selected product is no longer available for {original_name}."
                                    )
                    elif resolution_requested:
                        resolution_errors.append(
                            f"Select a product for '{original_name}' to continue."
                        )

                    if selected_product:
                        product_lookup[original_name] = selected_product
                        manual_mappings[original_name] = selected_product

                    unresolved_products.append(
                        {
                            "field": field_name,
                            "name": original_name,
                            "selected": selected_value or "",
                            "price": product_price_lookup.get(original_name),
                        }
                    )

                if not resolution_requested:
                    return render_template(
                        "events/upload_terminal_sales.html",
                        form=form,
                        event=ev,
                        open_locations=open_locations,
                        mapping_payload=payload,
                        mapping_filename=mapping_filename,
                        sales_summary=sales_summary,
                        sales_location_names=list(sales_summary.keys()),
                        default_mapping={
                            el.id: request.form.get(f"mapping-{el.id}", "")
                            for el in open_locations
                        },
                        unresolved_products=unresolved_products,
                        product_choices=product_choices,
                        resolution_errors=resolution_errors,
                        product_resolution_required=True,
                        price_discrepancies=price_discrepancies,
                        menu_mismatches=menu_mismatches,
                        warnings_required=warnings_required,
                        warnings_acknowledged=warnings_acknowledged,
                    )

                if (
                    len(manual_mappings) + len(pending_creations)
                    != len(unmatched_names)
                ):
                    resolution_errors.append(
                        "Select a product for each unmatched entry to continue."
                    )

                if resolution_errors:
                    return render_template(
                        "events/upload_terminal_sales.html",
                        form=form,
                        event=ev,
                        open_locations=open_locations,
                        mapping_payload=payload,
                        mapping_filename=mapping_filename,
                        sales_summary=sales_summary,
                        sales_location_names=list(sales_summary.keys()),
                        default_mapping={
                            el.id: request.form.get(f"mapping-{el.id}", "")
                            for el in open_locations
                        },
                        unresolved_products=unresolved_products,
                        product_choices=product_choices,
                        resolution_errors=resolution_errors,
                        product_resolution_required=True,
                        price_discrepancies=price_discrepancies,
                        menu_mismatches=menu_mismatches,
                        warnings_required=warnings_required,
                        warnings_acknowledged=warnings_acknowledged,
                    )

                for original_name in pending_creations:
                    existing_product = Product.query.filter_by(
                        name=original_name
                    ).first()
                    if existing_product:
                        product_lookup[original_name] = existing_product
                        manual_mappings[original_name] = existing_product
                        continue
                    price_value = product_price_lookup.get(original_name)
                    if price_value is None:
                        price_value = 0.0
                    new_product = Product(
                        name=original_name,
                        price=price_value,
                        cost=0.0,
                    )
                    db.session.add(new_product)
                    db.session.flush()
                    product_lookup[original_name] = new_product
                    manual_mappings[original_name] = new_product

                if manual_mappings and normalized_lookup:
                    for original_name, product in manual_mappings.items():
                        normalized = normalized_lookup.get(original_name, "")
                        if not normalized:
                            continue
                        alias = alias_lookup.get(normalized)
                        if alias is None:
                            alias = TerminalSaleProductAlias(
                                source_name=original_name,
                                normalized_name=normalized,
                                product=product,
                            )
                            db.session.add(alias)
                            alias_lookup[normalized] = alias
                        else:
                            alias.source_name = original_name
                            alias.product = product

                product_resolution_required = False

            pending_sales: list[dict] = []
            pending_totals: list[dict] = []
            selected_locations: list[str] = []
            issue_queue: list[dict] = []
            location_allowed_products: dict[int, set[int]] = {}
            for el in open_locations:
                selected_loc = request.form.get(f"mapping-{el.id}")
                if not selected_loc:
                    continue
                loc_sales = sales_summary.get(selected_loc)
                if not loc_sales:
                    continue
                location_updated = False
                price_issues: list[dict] = []
                menu_issues: list[dict] = []
                for prod_name, product_data in loc_sales["products"].items():
                    product = product_lookup.get(prod_name)
                    if not product:
                        continue
                    if el.location and product not in el.location.products:
                        el.location.products.append(product)
                    quantity_value = product_data.get("quantity", 0.0)
                    pending_sales.append(
                        {
                            "event_location_id": el.id,
                            "product_id": product.id,
                            "quantity": quantity_value,
                        }
                    )
                    location_updated = True

                    file_prices = product_data.get("prices") or []
                    if file_prices and not all(
                        _prices_match(price, product.price) for price in file_prices
                    ):
                        target_price = file_prices[0]
                        price_issues.append(
                            {
                                "product": product.name,
                                "product_id": product.id,
                                "file_prices": file_prices,
                                "app_price": product.price,
                                "sales_location": selected_loc,
                                "resolution": None,
                                "target_price": target_price,
                            }
                        )

                    allowed_products = location_allowed_products.get(el.id)
                    if allowed_products is None:
                        allowed_products = {
                            p.id for p in el.location.products
                        }
                        if el.location.current_menu is not None:
                            allowed_products.update(
                                p.id for p in el.location.current_menu.products
                            )
                        location_allowed_products[el.id] = allowed_products
                    if allowed_products and product.id not in allowed_products:
                        menu_issues.append(
                            {
                                "product": product.name,
                                "product_id": product.id,
                                "sales_location": selected_loc,
                                "menu_name": (
                                    el.location.current_menu.name
                                    if el.location.current_menu
                                    else None
                                ),
                                "resolution": None,
                            }
                        )
                if location_updated:
                    selected_locations.append(el.location.name)
                    pending_totals.append(
                        {
                            "event_location_id": el.id,
                            "source_location": selected_loc,
                            "total_quantity": loc_sales.get("total"),
                            "total_amount": loc_sales.get("total_amount"),
                        }
                    )
                if price_issues or menu_issues:
                    issue_queue.append(
                        {
                            "event_location_id": el.id,
                            "location_name": el.location.name,
                            "sales_location": selected_loc,
                            "price_issues": price_issues,
                            "menu_issues": menu_issues,
                        }
                    )

            if issue_queue:
                serializer = _terminal_sales_serializer()
                state_data = {
                    "queue": issue_queue,
                    "pending_sales": pending_sales,
                    "selected_locations": selected_locations,
                    "pending_totals": pending_totals,
                    "issue_index": 0,
                }
                state_token = serializer.dumps(state_data)
                return render_template(
                    "events/upload_terminal_sales.html",
                    form=form,
                    event=ev,
                    open_locations=open_locations,
                    mapping_payload=payload,
                    mapping_filename=mapping_filename,
                    sales_summary=sales_summary,
                    sales_location_names=list(sales_summary.keys()),
                    default_mapping={
                        el.id: request.form.get(f"mapping-{el.id}", "")
                        for el in open_locations
                    },
                    unresolved_products=[],
                    product_choices=product_choices,
                    resolution_errors=resolution_errors,
                    product_resolution_required=False,
                    price_discrepancies={},
                    menu_mismatches={},
                    warnings_required=False,
                    warnings_acknowledged=False,
                    state_token=state_token,
                    issue_index=0,
                    current_issue=issue_queue[0],
                    remaining_locations=len(issue_queue) - 1,
                    selected_locations=selected_locations,
                    issue_total=len(issue_queue),
                )

            updated_locations = _apply_pending_sales(pending_sales, pending_totals)
            if updated_locations:
                db.session.commit()
                log_activity(
                    "Uploaded terminal sales for event "
                    f"{event_id} from {mapping_filename or 'uploaded file'}"
                )
                flash(
                    "Terminal sales were imported for: "
                    + ", ".join(sorted(updated_locations)),
                    "success",
                )
            else:
                flash(
                    "No event locations were linked to the uploaded sales data.",
                    "warning",
                )
            return redirect(url_for("event.view_event", event_id=event_id))
        elif step:
            flash("Unable to process the uploaded sales data.", "danger")
            return redirect(url_for("event.upload_terminal_sales", event_id=event_id))

    if form.validate_on_submit():
        file = form.file.data
        filename = secure_filename(file.filename)
        ext = os.path.splitext(filename)[1].lower()
        filepath = os.path.join(current_app.config["UPLOAD_FOLDER"], filename)
        file.save(filepath)

        rows: list[dict] = []

        def add_row(
            loc,
            name,
            qty,
            price=None,
            amount=None,
            net_including_tax_total=None,
            discounts_total=None,
        ):
            if not loc or not isinstance(loc, str):
                return
            loc = loc.strip()
            if not loc:
                return
            if not name or not isinstance(name, str):
                return
            product_name = name.strip()
            if not product_name:
                return
            quantity_value = _to_float(qty)
            if quantity_value is None:
                return
            entry = {
                "location": loc,
                "product": product_name,
                "quantity": quantity_value,
            }
            price_value = _to_float(price)
            if price_value is not None:
                entry["price"] = price_value
            amount_value = _to_float(amount)
            if amount_value is not None:
                entry["amount"] = amount_value
            net_including_value = _to_float(net_including_tax_total)
            if net_including_value is not None:
                entry["net_including_tax_total"] = net_including_value
            discounts_value = _to_float(discounts_total)
            if discounts_value is not None:
                entry["discount_total"] = discounts_value
            rows.append(entry)

        try:
            if ext in {".xls", ".xlsx"}:
                def _iter_excel_rows(path: str, extension: str):
                    if extension == ".xls":
                        try:
                            import xlrd  # type: ignore
                        except ModuleNotFoundError:
                            try:
                                from app.vendor import xlrd  # type: ignore
                            except ImportError:
                                raise RuntimeError("legacy_xls_missing") from None
                        try:
                            book = xlrd.open_workbook(path)
                        except Exception as exc:  # pragma: no cover - defensive
                            raise RuntimeError("legacy_xls_error") from exc

                        try:
                            sheet = book.sheet_by_index(0)
                        except IndexError as exc:
                            raise RuntimeError("legacy_xls_error") from exc

                        try:
                            for row_idx in range(sheet.nrows):
                                yield [
                                    sheet.cell_value(row_idx, col_idx)
                                    for col_idx in range(sheet.ncols)
                                ]
                        finally:  # pragma: no branch - ensure resources freed
                            try:
                                book.release_resources()
                            except AttributeError:
                                pass
                    else:
                        try:
                            from openpyxl import load_workbook
                        except ImportError as exc:  # pragma: no cover - env issue
                            raise RuntimeError("xlsx_missing") from exc

                        try:
                            workbook = load_workbook(
                                path, read_only=True, data_only=True
                            )
                        except Exception as exc:
                            raise RuntimeError("xlsx_error") from exc

                        try:
                            sheet = workbook.active
                            for row in sheet.iter_rows(values_only=True):
                                yield list(row)
                        finally:
                            workbook.close()

                try:
                    rows_iter = _iter_excel_rows(filepath, ext)
                except RuntimeError as exc:
                    reason = str(exc)
                    if reason == "legacy_xls_missing":
                        flash(
                            "Legacy Excel support is unavailable on this server.",
                            "danger",
                        )
                        current_app.logger.exception(
                            "xlrd is required to read legacy .xls files"
                        )
                    else:
                        flash(
                            "The uploaded Excel file could not be read.",
                            "danger",
                        )
                        current_app.logger.exception(
                            "Failed to parse Excel file during terminal sales upload"
                        )
                    return redirect(
                        url_for("event.upload_terminal_sales", event_id=event_id)
                    )

                current_loc = None
                for row in rows_iter:
                    first = row[0] if row else None
                    second = row[1] if len(row) > 1 else None

                    if (second is None or second == "") and isinstance(first, str):
                        cleaned = first.strip()
                        if cleaned:
                            current_loc = cleaned
                        continue

                    if not current_loc or not isinstance(second, str):
                        continue

                    quantity = row[4] if len(row) > 4 else None
                    price = row[2] if len(row) > 2 else None
                    quantity_value = _to_float(quantity)
                    discounts = None
                    if quantity_value is not None and quantity_value != 0:
                        net_including = (
                            _to_float(row[7]) if len(row) > 7 else None
                        )
                        discounts = (
                            _to_float(row[8]) if len(row) > 8 else None
                        )
                        if net_including is not None:
                            price = (
                                net_including + (discounts or 0.0)
                            ) / quantity_value
                    amount = row[5] if len(row) > 5 else None
                    net_including_total = (
                        row[7] if len(row) > 7 else None
                    )
                    add_row(
                        current_loc,
                        second,
                        quantity,
                        price,
                        amount,
                        net_including_tax_total=net_including_total,
                        discounts_total=discounts,
                    )
            elif ext == ".pdf":
                import pdfplumber

                with pdfplumber.open(filepath) as pdf:
                    text = "\n".join(
                        page.extract_text() or "" for page in pdf.pages
                    )
                current_loc = None
                for line in text.splitlines():
                    line = line.strip()
                    if not line:
                        continue
                    if not line[0].isdigit():
                        current_loc = line
                        continue
                    if current_loc is None:
                        continue
                    parts = line.split()
                    idx = 1
                    while (
                        idx < len(parts)
                        and not parts[idx].replace(".", "", 1).isdigit()
                    ):
                        idx += 1
                    if idx + 2 < len(parts):
                        name = " ".join(parts[1:idx])
                        qty = parts[idx + 2]
                        add_row(current_loc, name, qty)
        finally:
            try:
                os.remove(filepath)
            except OSError:
                pass

        if not rows:
            flash(
                "No sales records were detected in the uploaded file.",
                "warning",
            )
        else:
            rows_data = rows

            sales_summary = _group_rows(rows_data)
            sales_location_names = list(sales_summary.keys())
            mapping_payload = json.dumps(
                {"rows": rows_data, "filename": filename}
            )
            mapping_filename = filename
            default_mapping = {
                el.id: el.location.name
                if el.location.name in sales_summary
                else ""
                for el in open_locations
            }

    return render_template(
        "events/upload_terminal_sales.html",
        form=form,
        event=ev,
        open_locations=open_locations,
        mapping_payload=mapping_payload,
        mapping_filename=mapping_filename,
        sales_summary=sales_summary,
        sales_location_names=sales_location_names,
        default_mapping=default_mapping,
        unresolved_products=unresolved_products,
        product_choices=product_choices,
        resolution_errors=resolution_errors,
        product_resolution_required=product_resolution_required,
        price_discrepancies=price_discrepancies,
        menu_mismatches=menu_mismatches,
        warnings_required=warnings_required,
        warnings_acknowledged=warnings_acknowledged,
    )


@event.route(
    "/events/<int:event_id>/locations/<int:el_id>/confirm",
    methods=["GET", "POST"],
)
@login_required
def confirm_location(event_id, el_id):
    el = db.session.get(EventLocation, el_id)
    if el is None or el.event_id != event_id:
        abort(404)
    form = EventLocationConfirmForm()
    if form.validate_on_submit():
        el.confirmed = True
        db.session.commit()
        log_activity(
            f"Confirmed event location {el_id} for event {event_id}"
        )
        flash("Location confirmed")
        return redirect(url_for("event.view_event", event_id=event_id))
    location, stand_items = _get_stand_items(el.location_id, event_id)
    stand_variances: list[dict] = []
    for entry in stand_items:
        sheet_values = entry.get("sheet_values")
        opening_val = getattr(sheet_values, "opening_count", None) or 0.0
        in_val = getattr(sheet_values, "transferred_in", None) or 0.0
        out_val = getattr(sheet_values, "transferred_out", None) or 0.0
        eaten_val = getattr(sheet_values, "eaten", None) or 0.0
        spoiled_val = getattr(sheet_values, "spoiled", None) or 0.0
        closing_val = getattr(sheet_values, "closing_count", None) or 0.0
        sales_val = entry.get("sales") or 0.0
        variance = (
            opening_val
            + in_val
            - out_val
            - sales_val
            - eaten_val
            - spoiled_val
            - closing_val
        )
        has_sheet = entry.get("sheet") is not None
        stand_variances.append(
            {
                "item": entry.get("item"),
                "report_unit_label": entry.get("report_unit_label"),
                "variance": variance if has_sheet else 0.0,
                "closing": closing_val if has_sheet else None,
            }
        )

    app_total_quantity = sum(float(sale.quantity or 0.0) for sale in el.terminal_sales)
    app_total_amount = sum(
        float(sale.quantity or 0.0) * float(sale.product.price or 0.0)
        for sale in el.terminal_sales
    )
    summary_record = el.terminal_sales_summary
    file_total_amount = None
    file_total_quantity = None
    source_location_name = None
    if summary_record is not None:
        file_total_amount = summary_record.total_amount
        file_total_quantity = summary_record.total_quantity
        source_location_name = summary_record.source_location
    amount_variance = None
    if file_total_amount is not None:
        amount_variance = app_total_amount - float(file_total_amount)
    return render_template(
        "events/confirm_location.html",
        form=form,
        event_location=el,
        stand_variances=stand_variances,
        sales_summary={
            "app_total_quantity": app_total_quantity,
            "app_total_amount": app_total_amount,
            "file_total_quantity": file_total_quantity,
            "file_total_amount": file_total_amount,
            "amount_variance": amount_variance,
            "source_location": source_location_name,
        },
        location=location,
    )


def _get_stand_items(location_id, event_id=None):
    location = db.session.get(Location, location_id)
    conversions = _conversion_mapping()
    stand_items = []
    seen = set()

    sales_by_item = {}
    sheet_map = {}
    if event_id is not None:
        el = EventLocation.query.filter_by(
            event_id=event_id,
            location_id=location_id,
        ).first()
        if el:
            for sale in el.terminal_sales:
                for ri in sale.product.recipe_items:
                    if ri.countable:
                        factor = ri.unit.factor if ri.unit else 1
                        sales_by_item[ri.item_id] = (
                            sales_by_item.get(ri.item_id, 0)
                            + sale.quantity * ri.quantity * factor
                        )
            for sheet in el.stand_sheet_items:
                sheet_map[sheet.item_id] = sheet

    for product_obj in location.products:
        for recipe_item in product_obj.recipe_items:
            if recipe_item.countable and recipe_item.item_id not in seen:
                seen.add(recipe_item.item_id)
                record = LocationStandItem.query.filter_by(
                    location_id=location_id,
                    item_id=recipe_item.item_id,
                ).first()
                expected = record.expected_count if record else 0
                sales = sales_by_item.get(recipe_item.item_id, 0)
                item = recipe_item.item
                recv_unit = next(
                    (u for u in item.units if u.receiving_default), None
                )
                trans_unit = next(
                    (u for u in item.units if u.transfer_default), None
                )
                stand_items.append(
                    _build_stand_item_entry(
                        item=item,
                        expected=expected,
                        sales=sales,
                        sheet=sheet_map.get(recipe_item.item_id),
                        recv_unit=recv_unit,
                        trans_unit=trans_unit,
                        conversions=conversions,
                    )
                )

    # Include any items directly assigned to the location that may not be
    # part of a product recipe (e.g. items received via purchase invoices).
    for record in LocationStandItem.query.filter_by(
        location_id=location_id
    ).all():
        if record.item_id in seen:
            continue
        item = record.item
        recv_unit = next((u for u in item.units if u.receiving_default), None)
        trans_unit = next((u for u in item.units if u.transfer_default), None)
        stand_items.append(
            _build_stand_item_entry(
                item=item,
                expected=record.expected_count,
                sales=sales_by_item.get(record.item_id, 0),
                sheet=sheet_map.get(record.item_id),
                recv_unit=recv_unit,
                trans_unit=trans_unit,
                conversions=conversions,
            )
        )
        seen.add(record.item_id)

    return location, stand_items


def build_sustainability_report(event_id: int) -> dict:
    """Aggregate waste, cost, and carbon metrics for an event."""

    carbon_per_unit = float(
        current_app.config.get("CARBON_EQ_PER_UNIT", 0.5)
    )
    query = (
        db.session.query(
            EventStandSheetItem,
            EventLocation,
            Location,
            Item,
        )
        .join(EventStandSheetItem.event_location)
        .join(EventLocation.location)
        .join(EventStandSheetItem.item)
        .filter(EventLocation.event_id == event_id)
    )

    totals = {"waste": 0.0, "cost": 0.0, "carbon": 0.0}
    location_totals = defaultdict(
        lambda: {"waste": 0.0, "cost": 0.0, "carbon": 0.0}
    )
    item_totals = defaultdict(lambda: {"waste": 0.0, "cost": 0.0, "carbon": 0.0})

    for sheet, _, location, item in query.all():
        eaten = sheet.eaten or 0.0
        spoiled = sheet.spoiled or 0.0
        waste_units = eaten + spoiled
        if waste_units == 0:
            continue

        unit_cost = item.cost or 0.0
        carbon_factor = getattr(item, "carbon_factor", None)
        if carbon_factor is None:
            carbon_factor = carbon_per_unit

        waste_cost = waste_units * unit_cost
        carbon_eq = waste_units * carbon_factor

        totals["waste"] += waste_units
        totals["cost"] += waste_cost
        totals["carbon"] += carbon_eq

        loc_bucket = location_totals[location.name]
        loc_bucket["waste"] += waste_units
        loc_bucket["cost"] += waste_cost
        loc_bucket["carbon"] += carbon_eq

        item_bucket = item_totals[item.name]
        item_bucket["waste"] += waste_units
        item_bucket["cost"] += waste_cost
        item_bucket["carbon"] += carbon_eq

    location_breakdown = [
        {
            "location": name,
            "waste": values["waste"],
            "cost": values["cost"],
            "carbon": values["carbon"],
        }
        for name, values in sorted(
            location_totals.items(), key=lambda item: item[1]["waste"], reverse=True
        )
    ]

    item_leaderboard = [
        {
            "item": name,
            "waste": values["waste"],
            "cost": values["cost"],
            "carbon": values["carbon"],
        }
        for name, values in sorted(
            item_totals.items(), key=lambda item: item[1]["waste"], reverse=True
        )
    ]

    goal_target = float(
        current_app.config.get("SUSTAINABILITY_WASTE_GOAL", 0) or 0.0
    )
    goal_progress = None
    goal_remaining = None
    goal_met = None
    if goal_target > 0:
        goal_remaining = max(goal_target - totals["waste"], 0.0)
        consumed_pct = min((totals["waste"] / goal_target) * 100, 100)
        goal_progress = round(100 - consumed_pct, 2)
        goal_met = totals["waste"] <= goal_target

    chart_data = {
        "labels": [entry["location"] for entry in location_breakdown],
        "datasets": [
            {
                "label": "Waste (units)",
                "backgroundColor": "#198754",
                "data": [entry["waste"] for entry in location_breakdown],
            },
            {
                "label": "Carbon (kg CO₂e)",
                "backgroundColor": "#0d6efd",
                "data": [entry["carbon"] for entry in location_breakdown],
            },
        ],
    }

    return {
        "totals": totals,
        "location_breakdown": location_breakdown,
        "item_leaderboard": item_leaderboard,
        "goal": {
            "target": goal_target if goal_target > 0 else None,
            "remaining": goal_remaining,
            "progress_pct": goal_progress,
            "met": goal_met,
        },
        "chart_data": chart_data,
    }


@event.route(
    "/events/<int:event_id>/stand_sheet/<int:location_id>",
    methods=["GET", "POST"],
)
@login_required
def stand_sheet(event_id, location_id):
    ev = db.session.get(Event, event_id)
    if ev is None:
        abort(404)
    el = EventLocation.query.filter_by(
        event_id=event_id, location_id=location_id
    ).first()
    if el is None:
        abort(404)
    if el.confirmed or ev.closed:
        flash(
            "This location is closed and the stand sheet cannot be modified."
        )
        return redirect(url_for("event.view_event", event_id=event_id))

    location, stand_items = _get_stand_items(location_id, event_id)

    if request.method == "POST":
        for entry in stand_items:
            item_id = entry["item"].id
            base_unit = entry.get("base_unit")
            report_unit = entry.get("report_unit") or base_unit
            sheet = EventStandSheetItem.query.filter_by(
                event_location_id=el.id,
                item_id=item_id,
            ).first()
            if not sheet:
                sheet = EventStandSheetItem(
                    event_location_id=el.id, item_id=item_id
                )
                db.session.add(sheet)
            opening = coerce_float(request.form.get(f"open_{item_id}"), default=0.0)
            transferred_in = coerce_float(
                request.form.get(f"in_{item_id}"), default=0.0
            )
            transferred_out = coerce_float(
                request.form.get(f"out_{item_id}"), default=0.0
            )
            eaten = coerce_float(request.form.get(f"eaten_{item_id}"), default=0.0)
            spoiled = coerce_float(
                request.form.get(f"spoiled_{item_id}"), default=0.0
            )
            closing = coerce_float(request.form.get(f"close_{item_id}"), default=0.0)
            sheet.opening_count = _convert_report_value_to_base(
                opening or 0, base_unit, report_unit
            )
            sheet.transferred_in = _convert_report_value_to_base(
                transferred_in or 0, base_unit, report_unit
            )
            sheet.transferred_out = _convert_report_value_to_base(
                transferred_out or 0, base_unit, report_unit
            )
            sheet.eaten = _convert_report_value_to_base(
                eaten or 0, base_unit, report_unit
            )
            sheet.spoiled = _convert_report_value_to_base(
                spoiled or 0, base_unit, report_unit
            )
            sheet.closing_count = _convert_report_value_to_base(
                closing or 0, base_unit, report_unit
            )
        db.session.commit()
        log_activity(
            f"Updated stand sheet for event {event_id} location {location_id}"
        )
        flash("Stand sheet saved")
        location, stand_items = _get_stand_items(location_id, event_id)

    return render_template(
        "events/stand_sheet.html",
        event=ev,
        location=location,
        stand_items=stand_items,
    )


@event.route("/events/<int:event_id>/sustainability")
@login_required
def sustainability_dashboard(event_id):
    ev = db.session.get(Event, event_id)
    if ev is None:
        abort(404)

    report = build_sustainability_report(event_id)
    chart_json = json.dumps(report["chart_data"])

    return render_template(
        "events/sustainability_dashboard.html",
        event=ev,
        report=report,
        chart_json=chart_json,
        print_view=False,
    )


@event.route("/events/<int:event_id>/sustainability/print")
@login_required
def sustainability_dashboard_print(event_id):
    ev = db.session.get(Event, event_id)
    if ev is None:
        abort(404)

    report = build_sustainability_report(event_id)
    chart_json = json.dumps(report["chart_data"])

    return render_template(
        "events/sustainability_dashboard.html",
        event=ev,
        report=report,
        chart_json=chart_json,
        print_view=True,
    )


@event.route("/events/<int:event_id>/sustainability/export.csv")
@login_required
def sustainability_dashboard_csv(event_id):
    ev = db.session.get(Event, event_id)
    if ev is None:
        abort(404)

    report = build_sustainability_report(event_id)

    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow(["Location", "Waste (units)", "Cost", "Carbon (kg CO₂e)"])
    for entry in report["location_breakdown"]:
        writer.writerow(
            [
                entry["location"],
                f"{entry['waste']:.2f}",
                f"{entry['cost']:.2f}",
                f"{entry['carbon']:.2f}",
            ]
        )

    writer.writerow([])
    writer.writerow(["Item", "Waste (units)", "Cost", "Carbon (kg CO₂e)"])
    for entry in report["item_leaderboard"]:
        writer.writerow(
            [
                entry["item"],
                f"{entry['waste']:.2f}",
                f"{entry['cost']:.2f}",
                f"{entry['carbon']:.2f}",
            ]
        )

    writer.writerow([])
    writer.writerow(["Totals", f"{report['totals']['waste']:.2f}", f"{report['totals']['cost']:.2f}", f"{report['totals']['carbon']:.2f}"])

    csv_response = make_response(output.getvalue())
    csv_response.headers["Content-Type"] = "text/csv"
    filename = f"sustainability-event-{event_id}.csv"
    csv_response.headers["Content-Disposition"] = f"attachment; filename={filename}"
    return csv_response


@event.route(
    "/events/<int:event_id>/count_sheet/<int:location_id>",
    methods=["GET", "POST"],
)
@login_required
def count_sheet(event_id, location_id):
    ev = db.session.get(Event, event_id)
    if ev is None:
        abort(404)
    el = EventLocation.query.filter_by(
        event_id=event_id, location_id=location_id
    ).first()
    if el is None:
        abort(404)
    if ev.closed:
        flash("This event is closed and cannot be modified.")
        return redirect(url_for("event.view_event", event_id=event_id))

    location, stand_items = _get_stand_items(location_id, event_id)

    if request.method == "POST":
        for entry in stand_items:
            item_id = entry["item"].id
            sheet = EventStandSheetItem.query.filter_by(
                event_location_id=el.id,
                item_id=item_id,
            ).first()
            if not sheet:
                sheet = EventStandSheetItem(
                    event_location_id=el.id, item_id=item_id
                )
                db.session.add(sheet)
            recv_qty = coerce_float(
                request.form.get(f"recv_{item_id}"), default=0.0
            ) or 0
            trans_qty = coerce_float(
                request.form.get(f"trans_{item_id}"), default=0.0
            ) or 0
            base_qty = coerce_float(
                request.form.get(f"base_{item_id}"), default=0.0
            ) or 0
            recv_factor = (
                entry["recv_unit"].factor if entry["recv_unit"] else 0
            )
            trans_factor = (
                entry["trans_unit"].factor if entry["trans_unit"] else 0
            )
            total = (
                recv_qty * recv_factor + trans_qty * trans_factor + base_qty
            )
            sheet.opening_count = recv_qty
            sheet.transferred_in = trans_qty
            sheet.transferred_out = base_qty
            sheet.closing_count = total
        db.session.commit()
        log_activity(
            f"Updated count sheet for event {event_id} location {location_id}"
        )
        flash("Count sheet saved")
        return redirect(url_for("event.view_event", event_id=event_id))

    return render_template(
        "events/count_sheet.html",
        event=ev,
        location=location,
        stand_items=stand_items,
    )


@event.route("/events/<int:event_id>/stand_sheets")
@login_required
def bulk_stand_sheets(event_id):
    ev = db.session.get(Event, event_id)
    if ev is None:
        abort(404)
    data = []
    for el in ev.locations:
        loc, items = _get_stand_items(el.location_id, event_id)
        data.append(
            {
                "location": loc,
                "stand_items": items,
            }
        )
    dt = datetime.now()
    generated_at_local = (
        f"{dt.month}/{dt.day}/{dt.year} {dt.strftime('%I:%M %p').lstrip('0')}"
    )
    return render_template(
        "events/bulk_stand_sheets.html",
        event=ev,
        data=data,
        generated_at_local=generated_at_local,
    )


@event.route("/events/<int:event_id>/count_sheets")
@login_required
def bulk_count_sheets(event_id):
    ev = db.session.get(Event, event_id)
    if ev is None:
        abort(404)
    data = []
    for el in ev.locations:
        loc, items = _get_stand_items(el.location_id, event_id)
        data.append(
            {
                "location": loc,
                "stand_items": items,
                "page_number": 1,
                "page_count": 1,
            }
        )
    return render_template(
        "events/bulk_count_sheets.html", event=ev, data=data
    )


@event.route("/events/<int:event_id>/close")
@login_required
def close_event(event_id):
    ev = db.session.get(Event, event_id)
    if ev is None:
        abort(404)
    for el in ev.locations:
        counted_item_ids = set()
        for sheet in el.stand_sheet_items:
            counted_item_ids.add(sheet.item_id)
            lsi = LocationStandItem.query.filter_by(
                location_id=el.location_id, item_id=sheet.item_id
            ).first()
            if not sheet.closing_count:
                if lsi:
                    db.session.delete(lsi)
                continue
            if not lsi:
                lsi = LocationStandItem(
                    location_id=el.location_id,
                    item_id=sheet.item_id,
                    purchase_gl_code_id=sheet.item.purchase_gl_code_id,
                )
                db.session.add(lsi)
            elif (
                lsi.purchase_gl_code_id is None
                and sheet.item.purchase_gl_code_id is not None
            ):
                lsi.purchase_gl_code_id = sheet.item.purchase_gl_code_id
            lsi.expected_count = sheet.closing_count

        if counted_item_ids:
            LocationStandItem.query.filter(
                LocationStandItem.location_id == el.location_id,
                ~LocationStandItem.item_id.in_(counted_item_ids),
            ).delete(synchronize_session=False)
        else:
            LocationStandItem.query.filter_by(
                location_id=el.location_id
            ).delete()

        TerminalSale.query.filter_by(event_location_id=el.id).delete()
    ev.closed = True
    db.session.commit()
    log_activity(f"Closed event {event_id}")
    flash("Event closed")
    return redirect(url_for("event.view_events"))


@event.route("/events/<int:event_id>/inventory_report")
@login_required
def inventory_report(event_id):
    """Display inventory variances and GL code totals for an event."""
    ev = db.session.get(Event, event_id)
    if ev is None:
        abort(404)

    rows = []
    gl_totals = {}
    grand_total = 0.0

    for el in ev.locations:
        loc = el.location
        for sheet in el.stand_sheet_items:
            item = sheet.item
            lsi = LocationStandItem.query.filter_by(
                location_id=loc.id, item_id=item.id
            ).first()
            expected = lsi.expected_count if lsi else 0
            variance = sheet.closing_count - expected
            cost_total = sheet.closing_count * item.cost
            gl_obj = item.purchase_gl_code_for_location(loc.id)
            gl_code = gl_obj.code if gl_obj else "Unassigned"
            rows.append(
                {
                    "location": loc,
                    "item": item,
                    "expected": expected,
                    "actual": sheet.closing_count,
                    "variance": variance,
                    "gl_code": gl_code,
                    "cost_total": cost_total,
                }
            )
            gl_totals[gl_code] = gl_totals.get(gl_code, 0.0) + cost_total
            grand_total += cost_total

    return render_template(
        "events/inventory_report.html",
        event=ev,
        rows=rows,
        gl_totals=gl_totals,
        grand_total=grand_total,
    )
